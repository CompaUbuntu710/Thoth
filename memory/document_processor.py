import os
import re
import threading
from datetime import datetime, timezone
from memory.vector_store import HAS_CHROMA, Settings

CHROMA_DIR = os.path.join(os.path.dirname(__file__), "chroma_db")

EXTENSION_MAP = {
    ".txt": "text",
    ".md": "markdown",
    ".csv": "csv",
    ".json": "json",
    ".py": "python",
    ".js": "javascript",
    ".ts": "typescript",
    ".html": "html",
    ".css": "css",
    ".yaml": "yaml",
    ".yml": "yaml",
    ".xml": "xml",
    ".sh": "shell",
    ".yml": "yaml",
}

UPLOAD_DIR = os.path.join(os.path.dirname(__file__), "..", "uploads")


def ensure_upload_dir():
    os.makedirs(UPLOAD_DIR, exist_ok=True)


def extract_text(file_path) -> str:
    ext = os.path.splitext(file_path)[1].lower()
    try:
        if ext == ".pdf":
            return _extract_pdf(file_path)
        elif ext == ".docx":
            return _extract_docx(file_path)
        elif ext == ".xlsx":
            return _extract_xlsx(file_path)
        else:
            return _extract_raw(file_path)
    except Exception as e:
        return f"[Error extrayendo texto: {e}]"


def _extract_pdf(path) -> str:
    from pypdf import PdfReader
    reader = PdfReader(path)
    pages = []
    for i, page in enumerate(reader.pages):
        text = page.extract_text() or ""
        pages.append(f"[Page {i+1}]\n{text.strip()}")
    return "\n\n".join(pages)


def _extract_docx(path) -> str:
    from docx import Document
    doc = Document(path)
    paras = [p.text.strip() for p in doc.paragraphs if p.text.strip()]
    return "\n\n".join(paras)


def _extract_xlsx(path) -> str:
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    parts = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        rows = []
        for row in ws.iter_row(values_only=True):
            vals = [str(v) if v is not None else "" for v in row]
            rows.append(" | ".join(vals))
        if rows:
            parts.append(f"[Sheet: {sheet}]\n" + "\n".join(rows))
    return "\n\n".join(parts)


def _extract_raw(path) -> str:
    try:
        with open(path, "r", encoding="utf-8") as f:
            return f.read()
    except UnicodeDecodeError:
        with open(path, "r", encoding="latin-1") as f:
            return f.read()


def chunk_text(text: str, chunk_size=800, overlap=100) -> list[str]:
    if not text.strip():
        return []
    paragraphs = re.split(r"\n\s*\n", text.strip())
    chunks = []
    current = ""
    for para in paragraphs:
        para = para.strip()
        if not para:
            continue
        if len(current) + len(para) + 2 > chunk_size and current:
            chunks.append(current.strip())
            current = current[-overlap:] if overlap > 0 else ""
        current = (current + "\n\n" + para).strip()
    if current.strip():
        chunks.append(current.strip())
    return chunks if chunks else [text.strip()[:chunk_size]]


class DocumentStore:
    """Almacén de documentos con extracción de texto, chunking y ChromaDB."""

    def __init__(self, path=None):
        self._lock = threading.Lock()
        self.path = path or CHROMA_DIR
        self._client = None
        self._collection = None

    @property
    def client(self):
        if self._client is None:
            os.makedirs(self.path, exist_ok=True)
            import chromadb
            self._client = chromadb.PersistentClient(
                path=self.path,
                settings=Settings(anonymized_telemetry=False),
            )
        return self._client

    @property
    def collection(self):
        if self._collection is None:
            self._collection = self.client.get_or_create_collection(
                name="thoth_documents",
                metadata={"hnsw:space": "cosine"},
            )
        return self._collection

    def index_file(self, file_path: str, original_name: str = None) -> dict:
        if not HAS_CHROMA:
            return {"status": "error", "message": "ChromaDB no disponible"}
        name = original_name or os.path.basename(file_path)
        ext = os.path.splitext(name)[1].lower()
        text = extract_text(file_path)
        if not text.strip():
            return {"status": "error", "message": "No se pudo extraer texto"}
        chunks = chunk_text(text)
        if not chunks:
            return {"status": "error", "message": "Texto vacío después de chunking"}
        now = datetime.now(timezone.utc).isoformat()
        ids = [f"{name}::{i}" for i in range(len(chunks))]
        metadatas = [{
            "file_name": name,
            "file_type": EXTENSION_MAP.get(ext, ext.lstrip(".") if ext else "unknown"),
            "chunk_index": i,
            "total_chunks": len(chunks),
            "created_at": now,
            "type": "document",
        } for i in range(len(chunks))]
        with self._lock:
            try:
                self.collection.add(
                    ids=ids,
                    documents=chunks,
                    metadatas=metadatas,
                )
            except Exception as e:
                return {"status": "error", "message": str(e)}
        return {
            "status": "ok",
            "file_name": name,
            "chunks": len(chunks),
            "characters": len(text),
        }

    def search(self, query: str, n_results=5, file_filter: str = None) -> list[dict]:
        if not HAS_CHROMA or not query:
            return []
        try:
            where = {"type": "document"}
            if file_filter:
                where["file_name"] = file_filter
            with self._lock:
                results = self.collection.query(
                    query_texts=[query],
                    n_results=n_results,
                    where=where,
                )
            items = []
            if results["documents"] and results["documents"][0]:
                for i, doc in enumerate(results["documents"][0]):
                    meta = results["metadatas"][0][i] if results["metadatas"] else {}
                    items.append({
                        "text": doc,
                        "file_name": meta.get("file_name", ""),
                        "file_type": meta.get("file_type", ""),
                        "chunk_index": meta.get("chunk_index", 0),
                        "total_chunks": meta.get("total_chunks", 1),
                        "score": results["distances"][0][i] if results["distances"] else 0,
                    })
            return items
        except Exception:
            return []

    def list_documents(self) -> list[dict]:
        if not HAS_CHROMA:
            return []
        try:
            seen = {}
            with self._lock:
                all_meta = self.collection.get(include=["metadatas"])
                if not all_meta or not all_meta["metadatas"]:
                    return []
                for m in all_meta["metadatas"]:
                    fn = m.get("file_name", "unknown")
                    if fn not in seen:
                        seen[fn] = {
                            "file_name": fn,
                            "file_type": m.get("file_type", ""),
                            "total_chunks": m.get("total_chunks", 1),
                            "created_at": m.get("created_at", ""),
                        }
            return list(seen.values())
        except Exception:
            return []

    def delete_file(self, file_name: str) -> bool:
        if not HAS_CHROMA:
            return False
        try:
            with self._lock:
                existing = self.collection.get(
                    where={"file_name": file_name},
                    include=[],
                )
                if existing and existing["ids"]:
                    self.collection.delete(ids=existing["ids"])
            return True
        except Exception:
            return False

    def count(self) -> int:
        if not HAS_CHROMA:
            return 0
        try:
            return self.collection.count()
        except Exception:
            return 0


document_store = DocumentStore()
ensure_upload_dir()
